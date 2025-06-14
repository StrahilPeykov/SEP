from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase
from openpyxl import load_workbook

from core.tests.helpers.get_stream_bytes import get_stream_bytes
from core.tests.setup_functions import tech_companies_setup
from core.models import (
    ProductionEnergyEmission,
    ProductionEnergyEmissionReference,
    ProductionEnergyEmissionReferenceFactor,
)


class ProductionEnergyEmissionExportImportTests(APITestCase):
    def setUp(self):
        tech_companies_setup(self)

        # create a reference with one factor
        self.reference = ProductionEnergyEmissionReference.objects.create(common_name="ProdRef")
        ProductionEnergyEmissionReferenceFactor.objects.create(
            emission_reference=self.reference,
            lifecycle_stage="OTHER",
            co_2_emission_factor_biogenic=0.3,
            co_2_emission_factor_non_biogenic=0.7,
        )

        # attach one production energy emission to self.iphone
        self.emission = ProductionEnergyEmission.objects.create(
            parent_product=self.iphone,
            energy_consumption=123.4,
            reference=self.reference,
        )

    def test_export_csv(self):
        """
        Test exporting production energy emissions to CSV format.
        """
        url = reverse("product-production-energy-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv")
        cd = response["Content-Disposition"]
        self.assertIn(self.apple.name, cd)
        self.assertIn(".csv", cd)

        raw = get_stream_bytes(response)
        header = raw.decode("utf-8").splitlines()[0]
        self.assertIn("reference", header)
        self.assertIn("energy_consumption", header)

    def test_export_csv_template(self):
        """
        Test exporting a template for production energy emissions in CSV format.
        """
        url = reverse("product-production-energy-emissions-export-csv", args=[self.apple.id, self.iphone.id]) + "?template=true"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        raw = get_stream_bytes(response)
        header = raw.decode("utf-8").splitlines()[0]
        self.assertIn("reference", header)
        self.assertIn("energy_consumption", header)

    def test_export_xlsx(self):
        """
        Test exporting production energy emissions to XLSX format.
        """
        url = reverse("product-production-energy-emissions-export-xlsx", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        cd = response["Content-Disposition"]
        self.assertIn(self.apple.name, cd)
        self.assertIn(".xlsx", cd)

        raw = get_stream_bytes(response)
        wb = load_workbook(filename=BytesIO(raw), read_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.rows)]
        self.assertIn("reference", headers)
        self.assertIn("energy_consumption", headers)

    def test_export_xlsx_template(self):
        """
        Test exporting a template for production energy emissions in XLSX format.
        """
        url = reverse("product-production-energy-emissions-export-xlsx", args=[self.apple.id, self.iphone.id]) + "?template=true"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        raw = get_stream_bytes(response)
        wb = load_workbook(filename=BytesIO(raw), read_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.rows)]
        self.assertIn("reference", headers)
        self.assertIn("energy_consumption", headers)

    def test_export_csv_not_logged_in(self):
        """
        Test exporting production energy emissions to CSV format when not logged in.
        """
        self.client.credentials()
        url = reverse("product-production-energy-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_export_xlsx_not_logged_in(self):
        """
        Test exporting production energy emissions to XLSX format when not logged in.
        """
        self.client.credentials()
        url = reverse("product-production-energy-emissions-export-xlsx", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_import_csv_creates(self):
        """
        Test importing production energy emissions from a CSV file that creates new entries.
        """
        url_exp = reverse("product-production-energy-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        raw = get_stream_bytes(self.client.get(url_exp))

        # count existing
        existing_count = ProductionEnergyEmission.objects.filter(parent_product=self.iphone).count()
        self.assertGreater(existing_count, 0)

        ProductionEnergyEmission.objects.filter(parent_product=self.iphone).delete()
        self.assertEqual(ProductionEnergyEmission.objects.filter(parent_product=self.iphone).count(), 0)

        url_imp = reverse("product-production-energy-emissions-import-tabular", args=[self.apple.id, self.iphone.id])
        uploaded = SimpleUploadedFile("prod.csv", raw, content_type="text/csv")
        resp_imp = self.client.post(url_imp, {"file": uploaded}, format="multipart")

        self.assertEqual(resp_imp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(resp_imp.data), existing_count)
        self.assertEqual(ProductionEnergyEmission.objects.filter(parent_product=self.iphone).count(), existing_count)

    def test_import_csv_fuzzy(self):
        """
        Test importing production energy emissions with a CSV file that has fuzzy matching.
        """
        url_exp = reverse("product-production-energy-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        raw = get_stream_bytes(self.client.get(url_exp))

        raw = raw.replace(b"ProdRef", b"PrdRef")  # fuzzy tweak

        # count existing
        existing_count = ProductionEnergyEmission.objects.filter(parent_product=self.iphone).count()

        ProductionEnergyEmission.objects.filter(parent_product=self.iphone).delete()
        url_imp = reverse("product-production-energy-emissions-import-tabular", args=[self.apple.id, self.iphone.id])
        uploaded = SimpleUploadedFile("prod_fuzzy.csv", raw, content_type="text/csv")
        resp_imp = self.client.post(url_imp, {"file": uploaded}, format="multipart")

        self.assertEqual(resp_imp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(resp_imp.data), existing_count)
        self.assertEqual(ProductionEnergyEmission.objects.filter(parent_product=self.iphone).count(), existing_count)

    def test_import_wrong_extension(self):
        """
        Test importing production energy emissions with a file that has the wrong extension.
        """
        url = reverse("product-production-energy-emissions-import-tabular",
                      args=[self.apple.id, self.iphone.id])
        bad_file = SimpleUploadedFile("bad.txt", b"Some content", content_type="text/plain")
        resp = self.client.post(url, {"file": bad_file}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)

    def test_import_wrong_key(self):
        """
        Test importing production energy emissions with a file that has the wrong key.
        """
        url = reverse("product-production-energy-emissions-import-tabular",
                      args=[self.apple.id, self.iphone.id])
        bad_file = SimpleUploadedFile("bad.csv", b"Some content", content_type="text/csv")
        resp = self.client.post(url, {"wrong_key": bad_file}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)