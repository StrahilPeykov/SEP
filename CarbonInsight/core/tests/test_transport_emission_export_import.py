from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase
from openpyxl import load_workbook

from core.tests.helpers.get_stream_bytes import get_stream_bytes
from core.tests.setup_functions import tech_companies_setup
from core.models import TransportEmission, TransportEmissionReference, TransportEmissionReferenceFactor


class TransportEmissionExportImportTests(APITestCase):
    def setUp(self):
        tech_companies_setup(self)
        # create a reference with one factor
        self.reference = TransportEmissionReference.objects.create(common_name="TestRef")
        TransportEmissionReferenceFactor.objects.create(
            emission_reference=self.reference,
            lifecycle_stage="OTHER",
            co_2_emission_factor_biogenic=0.5,
            co_2_emission_factor_non_biogenic=1.0,
        )
        # attach one transport emission to self.iphone
        self.emission = TransportEmission.objects.create(
            parent_product=self.iphone,
            distance=100.0,
            weight=10.0,
            reference=self.reference,
        )

    def test_export_csv(self):
        """
        Test exporting transport emissions to CSV format.
        """
        url = reverse("product-transport-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "text/csv")
        cd = response["Content-Disposition"]
        self.assertIn(self.apple.name, cd)
        self.assertIn(".csv", cd)

        raw = get_stream_bytes(response)
        text = raw.decode("utf-8")
        lines = text.splitlines()
        self.assertGreater(len(lines), 1, "Should have header + data row")
        self.assertIn("distance", lines[0])
        self.assertIn("weight", lines[0])

    def test_export_csv_template(self):
        """
        Test exporting a template for transport emissions in CSV format.
        """
        url = reverse("product-transport-emissions-export-csv", args=[self.apple.id, self.iphone.id]) + "?template=true"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        raw = get_stream_bytes(response)
        lines = raw.decode("utf-8").splitlines()
        self.assertEqual(len(lines), 1, "Template should only have header row")

    def test_export_xlsx(self):
        """
        Test exporting transport emissions to XLSX format.
        """
        url = reverse("product-transport-emissions-export-xlsx", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        cd = response["Content-Disposition"]
        self.assertIn(self.apple.name, cd)
        self.assertIn(".xlsx", cd)

        raw = get_stream_bytes(response)
        wb = load_workbook(filename=BytesIO(raw), read_only=True)
        sheet = wb.active
        rows = list(sheet.rows)
        self.assertGreater(len(rows), 1)
        headers = [c.value for c in rows[0]]
        self.assertIn("distance", headers)
        self.assertIn("weight", headers)

    def test_export_xlsx_template(self):
        """
        Test exporting a template for transport emissions in XLSX format.
        """
        url = reverse("product-transport-emissions-export-xlsx", args=[self.apple.id, self.iphone.id]) + "?template=true"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        raw = get_stream_bytes(response)
        wb = load_workbook(filename=BytesIO(raw), read_only=True)
        sheet = wb.active
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 1, "Template should only have header row")

    def test_export_csv_not_logged_in(self):
        """
        Test that exporting transport emissions to CSV format requires authentication.
        """
        self.client.credentials()  # remove auth
        url = reverse("product-transport-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_export_xlsx_not_logged_in(self):
        """
        Test that exporting transport emissions to XLSX format requires authentication.
        """
        self.client.credentials()
        url = reverse("product-transport-emissions-export-xlsx", args=[self.apple.id, self.iphone.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_import_csv_creates(self):
        """
        Test importing transport emissions from CSV format.
        """
        # export first
        url_exp = reverse("product-transport-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        resp_exp = self.client.get(url_exp)
        raw = get_stream_bytes(resp_exp)

        # count existing
        existing_count = TransportEmission.objects.filter(parent_product=self.iphone).count()

        # delete existing
        TransportEmission.objects.filter(parent_product=self.iphone).delete()
        self.assertEqual(TransportEmission.objects.filter(parent_product=self.iphone).count(), 0)

        # import
        url_imp = reverse("product-transport-emissions-import-tabular", args=[self.apple.id, self.iphone.id])
        uploaded = SimpleUploadedFile("trans.csv", raw, content_type="text/csv")
        resp_imp = self.client.post(url_imp, {"file": uploaded}, format="multipart")

        self.assertEqual(resp_imp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(resp_imp.data), existing_count)
        self.assertEqual(TransportEmission.objects.filter(parent_product=self.iphone).count(), existing_count)

    def test_import_csv_fuzzy(self):
        """
        Test importing transport emissions from CSV format with fuzzy matching.
        """
        url_exp = reverse("product-transport-emissions-export-csv", args=[self.apple.id, self.iphone.id])
        raw = get_stream_bytes(self.client.get(url_exp))

        # tweak reference name in CSV for fuzzy matching
        raw = raw.replace(b"TestRef", b"TstRef")

        # count existing
        existing_count = TransportEmission.objects.filter(parent_product=self.iphone).count()

        TransportEmission.objects.filter(parent_product=self.iphone).delete()
        url_imp = reverse("product-transport-emissions-import-tabular", args=[self.apple.id, self.iphone.id])
        uploaded = SimpleUploadedFile("trans_fuzzy.csv", raw, content_type="text/csv")
        resp_imp = self.client.post(url_imp, {"file": uploaded}, format="multipart")

        self.assertEqual(resp_imp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(resp_imp.data), existing_count)
        self.assertEqual(TransportEmission.objects.filter(parent_product=self.iphone).count(), existing_count)

    def test_import_wrong_extension(self):
        """
        Test importing transport emissions with a file that has the wrong extension.
        """
        url = reverse("product-transport-emissions-import-tabular",
                      args=[self.apple.id, self.iphone.id])
        bad_file = SimpleUploadedFile("bad.txt", b"Some content", content_type="text/plain")
        resp = self.client.post(url, {"file": bad_file}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)

    def test_import_wrong_key(self):
        """
        Test importing transport emissions with a file that has the wrong key in the form data.
        """
        url = reverse("product-transport-emissions-import-tabular",
                      args=[self.apple.id, self.iphone.id])
        bad_file = SimpleUploadedFile("bad.csv", b"Some content", content_type="text/csv")
        resp = self.client.post(url, {"wrong_key": bad_file}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)